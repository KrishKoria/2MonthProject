"""Fetch publicly-available CMS/OIG source material.

Best-effort downloads. Each source is independent: if one fails, others continue.
Anything that fails is covered by generate_synthetic_corpus.py afterward.

Targets:
- CMS Medicare Claims Processing Manual (Pub. 100-04), selected chapters (PDF → text)
- OIG compliance/fraud guidance documents (PDF → text)
- NCCI Practitioner PTP Edits (ZIP/XLSX → CSV)

PDFs and ZIPs are kept under data/raw/ (gitignored). Extracted text lands under
data/policy_docs/<topic>/*.txt. NCCI ends up at data/ncci/practitioner_ptp_edits.csv.

Usage:
    uv run python -m scripts.fetch_public_data
"""

from __future__ import annotations

import io
import logging
import re
import sys
import zipfile
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import settings

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s - %(message)s")
logger = logging.getLogger("fetch_public_data")

HTTP_TIMEOUT = 60
HEADERS = {
    "User-Agent": "claims-investigation-assistant/0.1 (research demo; +local)",
    "Accept": "*/*",
}

RAW_DIR = settings.DATA_DIR / "raw" / "public_sources"
NCCI_CSV = settings.ncci_dir / "practitioner_ptp_edits.csv"

# Candidate URLs — tried in order, first HTTP 200 wins.
CMS_MANUAL_CHAPTERS: dict[str, list[str]] = {
    "clm104c01_general_billing": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c01.pdf",
    ],
    "clm104c12_physicians_em": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c12.pdf",
    ],
    "clm104c13_radiology": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c13.pdf",
    ],
    "clm104c16_laboratory": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c16.pdf",
    ],
    "clm104c17_drugs_biologicals": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c17.pdf",
    ],
    "clm104c23_fee_schedule_coding": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c23.pdf",
    ],
    "clm104c26_cms1500_form": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c26.pdf",
    ],
    "clm104c30_financial_liability": [
        "https://www.cms.gov/Regulations-and-Guidance/Guidance/Manuals/Downloads/clm104c30.pdf",
    ],
}

OIG_CANDIDATE_URLS = [
    "https://oig.hhs.gov/documents/compliance-guidance/1135/hhs-oig-gcpg-2023.pdf",
    "https://oig.hhs.gov/documents/root/804/physician.pdf",
    "https://oig.hhs.gov/documents/root/803/hospital.pdf",
]

NCCI_LANDING = (
    "https://www.cms.gov/medicare/coding-billing/"
    "national-correct-coding-initiative-ncci-edits"
)


# ---------------------------------------------------------------------------
# HTTP helpers


def _get(url: str, stream: bool = False) -> requests.Response | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=HTTP_TIMEOUT, stream=stream)
        if r.status_code == 200:
            return r
        logger.warning("GET %s -> HTTP %d", url, r.status_code)
    except requests.RequestException as exc:
        logger.warning("GET %s failed: %s", url, exc)
    return None


def _download(url: str, dest: Path) -> bool:
    dest.parent.mkdir(parents=True, exist_ok=True)
    r = _get(url, stream=True)
    if r is None:
        return False
    with dest.open("wb") as f:
        for chunk in r.iter_content(chunk_size=65536):
            if chunk:
                f.write(chunk)
    logger.info("Saved %s (%d bytes) from %s", dest, dest.stat().st_size, url)
    return True


# ---------------------------------------------------------------------------
# PDF → text


def _pdf_to_text(pdf_path: Path) -> str:
    try:
        reader = PdfReader(str(pdf_path))
    except Exception as exc:
        logger.warning("Could not open PDF %s: %s", pdf_path, exc)
        return ""
    parts: list[str] = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:  # pragma: no cover — pypdf edge cases
            continue
    text = "\n".join(parts)
    # Light cleanup: collapse excessive whitespace
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ---------------------------------------------------------------------------
# Source handlers


def fetch_cms_manual() -> int:
    """Download selected CMS manual chapters, extract text. Return chapters fetched."""
    out_dir = settings.policy_docs_dir / "cms_claims_manual"
    out_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for slug, candidates in CMS_MANUAL_CHAPTERS.items():
        pdf_path = RAW_DIR / "cms_claims_manual" / f"{slug}.pdf"
        if not pdf_path.exists():
            ok = False
            for url in candidates:
                if _download(url, pdf_path):
                    ok = True
                    break
            if not ok:
                logger.warning("CMS manual chapter %s unavailable", slug)
                continue
        text = _pdf_to_text(pdf_path)
        if not text or len(text) < 5000:
            logger.warning("Chapter %s extracted text suspiciously short (%d chars)", slug, len(text))
            continue
        header = f"Source: CMS Medicare Claims Processing Manual\nChapter: {slug}\n\n"
        (out_dir / f"{slug}.txt").write_text(header + text, encoding="utf-8")
        count += 1
    logger.info("CMS manual chapters written: %d", count)
    return count


def fetch_oig_guidance() -> int:
    """Download canonical OIG compliance/fraud PDFs, extract text."""
    out_dir = settings.policy_docs_dir / "fraud_guidelines"
    out_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for url in OIG_CANDIDATE_URLS:
        slug = Path(url).stem
        pdf_path = RAW_DIR / "oig" / f"{slug}.pdf"
        if not pdf_path.exists() and not _download(url, pdf_path):
            continue
        text = _pdf_to_text(pdf_path)
        if not text or len(text) < 3000:
            continue
        header = f"Source: OIG ({url})\n\n"
        (out_dir / f"{slug}.txt").write_text(header + text, encoding="utf-8")
        count += 1
    logger.info("OIG guidance documents written: %d", count)
    return count


def _find_ncci_zip_url() -> str | None:
    """Scrape the NCCI landing page for the latest practitioner PTP edits ZIP."""
    r = _get(NCCI_LANDING)
    if r is None:
        return None
    soup = BeautifulSoup(r.text, "lxml")
    candidates: list[str] = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        low = href.lower()
        if ".zip" in low and "practitioner" in low and ("ptp" in low or "edit" in low):
            candidates.append(urljoin(NCCI_LANDING, href))
    if not candidates:
        # Fall back to any ZIP on the page mentioning "practitioner"
        for a in soup.find_all("a", href=True):
            href = a["href"]
            low = href.lower()
            if ".zip" in low and "practitioner" in low:
                candidates.append(urljoin(NCCI_LANDING, href))
    # Sort by any embedded 4-digit year in the URL, newest first
    def year_key(u: str) -> int:
        m = re.search(r"(20\d{2})", u)
        return int(m.group(1)) if m else 0
    candidates.sort(key=year_key, reverse=True)
    return candidates[0] if candidates else None


def fetch_ncci() -> bool:
    """Download NCCI practitioner PTP edits ZIP, extract first xlsx, normalize to CSV."""
    zip_url = _find_ncci_zip_url()
    if not zip_url:
        logger.warning("Could not discover NCCI ZIP on %s", NCCI_LANDING)
        return False
    zip_path = RAW_DIR / "ncci" / Path(zip_url).name
    if not zip_path.exists() and not _download(zip_url, zip_path):
        return False
    try:
        with zipfile.ZipFile(zip_path) as zf:
            xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
            if not xlsx_names:
                logger.warning("NCCI ZIP %s contains no xlsx", zip_path)
                return False
            with zf.open(xlsx_names[0]) as xf:
                data = xf.read()
    except zipfile.BadZipFile as exc:
        logger.warning("Bad NCCI ZIP %s: %s", zip_path, exc)
        return False

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return False

    # CMS NCCI header row contains "Column 1", "Column 2", "Effective Date",
    # "Deletion Date", and a modifier indicator column. Header may be preceded
    # by a title row. Find the header by searching for "Column 1".
    header_idx = None
    for i, row in enumerate(rows[:20]):
        if row and any(isinstance(c, str) and "column 1" in c.lower() for c in row):
            header_idx = i
            break
    if header_idx is None:
        logger.warning("Could not locate NCCI header row")
        return False

    header = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]

    def col(*needles: str) -> int | None:
        for idx, name in enumerate(header):
            if all(n in name for n in needles):
                return idx
        return None

    c1 = col("column 1")
    c2 = col("column 2")
    eff = col("effective", "date")
    deld = col("deletion", "date")
    mod = col("modifier")

    if None in (c1, c2, eff):
        logger.warning("NCCI header missing required columns: %s", header)
        return False

    NCCI_CSV.parent.mkdir(parents=True, exist_ok=True)
    import csv

    written = 0
    with NCCI_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["code_1", "code_2", "effective_date", "deletion_date", "modifier_indicator"])
        for row in rows[header_idx + 1 :]:
            if not row or row[c1] is None or row[c2] is None:
                continue
            code_1 = str(row[c1]).strip()
            code_2 = str(row[c2]).strip()
            if not code_1 or not code_2:
                continue
            effective = row[eff]
            deletion = row[deld] if deld is not None else None
            modi = row[mod] if mod is not None else None
            # Normalize code pair (sorted tuple for lookup invariance)
            pair = tuple(sorted((code_1, code_2)))
            w.writerow([
                pair[0],
                pair[1],
                _as_date(effective),
                _as_date(deletion),
                "" if modi is None else str(modi).strip(),
            ])
            written += 1
    logger.info("NCCI CSV written: %d edit rows at %s", written, NCCI_CSV)
    return written > 0


def _as_date(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return v.strftime("%Y-%m-%d")  # datetime
    except AttributeError:
        s = str(v).strip()
        # Common CMS format: "20210101" or "1/1/2021"
        m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
        if m:
            return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        return s


# ---------------------------------------------------------------------------
# Main


def main() -> int:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    results = {
        "cms_manual_chapters": fetch_cms_manual(),
        "oig_documents": fetch_oig_guidance(),
        "ncci_rows": fetch_ncci(),
    }
    logger.info("Fetch results: %s", results)
    # Non-fatal: synthetic supplement will fill gaps
    return 0


if __name__ == "__main__":
    sys.exit(main())
